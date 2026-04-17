import pandas as pd
from openpyxl import load_workbook
import os

# DATABASE INTEGRATO CON I COMUNI MANCANTI SEGNALATI
db_comuni = {
    "castelgomberto": "VI", "san vito di leguzzano": "VI", "arcugnano": "VI", "marano vicentino": "VI",
    "carrè": "VI", "breganze": "VI", "santorso": "VI", "costabissara": "VI", "cappella maggiore": "TV",
    "caldogno": "VI", "marostica": "VI", "sossano": "VI", "montegalda": "VI", "longare": "VI",
    "dueville": "VI", "creazzo": "VI", "camisano": "VI", "pojana": "VI", "tezze": "VI", "recoaro": "VI",
    "isola vicentina": "VI", "malo": "VI", "schio": "VI", "bassano": "VI", "thiene": "VI", "vicenza": "VI",
    "montecchio": "VI", "arzignano": "VI", "valdagno": "VI", "sandrigo": "VI", "bolzano vicentino": "VI",
    "altavilla": "VI", "sovizzo": "VI", "quinto": "VI", "monticello": "VI", "caltrano": "VI", "zugliano": "VI",
    "asiago": "VI", "roana": "VI", "gallio": "VI", "eraclea": "VE", "jesolo": "VE", "padova": "PD", "verona": "VR",
    "treviso": "TV", "rovigo": "RO", "belluno": "BL", "modena": "MO", "bergamo": "BG", "mestre": "VE"
}

def completa():
    file_path = "web/Pipistrelli 2025.xlsx"
    if not os.path.exists(file_path): return

    print(f"--- COMPLETAMENTO FINALE FISICO EXCEL ---")
    wb = load_workbook(file_path)
    ws = wb.active

    col_comune = col_prov = col_loc = -1
    for r in range(1, 10):
        for cell in ws[r]:
            val = str(cell.value).lower() if cell.value else ""
            if "comune" in val: col_comune = cell.column
            if "prov" in val: col_prov = cell.column
            if "localit" in val: col_loc = cell.column
        if col_prov != -1: break

    modificati = 0
    for row in range(2, ws.max_row + 1):
        comune_val = str(ws.cell(row=row, column=col_comune).value or "").lower()
        localita_val = str(ws.cell(row=row, column=col_loc).value or "").lower() if col_loc != -1 else ""
        prov_cell = ws.cell(row=row_idx if 'row_idx' in locals() else row, column=col_prov)
        # Fix per variabile row_idx errata nel loop precedente
        prov_cell = ws.cell(row=row, column=col_prov)
        prov_val = str(prov_cell.value or "").strip()

        if prov_val in ["", "None", "nan", "-", "0"]:
            search_text = comune_val + " " + localita_val
            if not search_text.strip(): continue # Salta righe vuote

            sigla_trovata = None
            for nome, sigla in db_comuni.items():
                if nome in search_text:
                    sigla_trovata = sigla
                    break

            if sigla_trovata:
                prov_cell.value = sigla_trovata
                modificati += 1

    wb.save(file_path)
    print(f"✅ OPERAZIONE CONCLUSA: {modificati} province inserite fisicamente.")

if __name__ == "__main__":
    completa()
