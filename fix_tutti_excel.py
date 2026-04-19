import json
import os
import shutil
import math
from openpyxl import load_workbook

def get_distance(lat1, lon1, lat2, lon2):
    return math.sqrt((lat1 - lat2)**2 + (lon1 - lon2)**2)

def fix_excel_files():
    years = [2022, 2023, 2024, 2025]
    json_path = 'app/src/main/assets/comuni_italiani.json'

    if not os.path.exists(json_path):
        print(f"Errore: {json_path} non trovato.")
        return

    with open(json_path, encoding='utf-8') as f:
        db_comuni = json.load(f)

    # Dizionario di correzione manuale potenziato
    CORREZIONI_NOMI = {
        "due ville": "dueville",
        "dueville": "dueville",
        "montecchio": "montecchio maggiore",
        "montecchio vicenza": "montecchio maggiore",
        "montecchio maggiore": "montecchio maggiore",
        "montecchio precalcino": "montecchio precalcino",
        "altavilla": "altavilla vicentina",
        "quinto": "quinto vicentino",
        "camisano": "camisano vicentino",
        "camisano vicentino": "camisano vicentino",
        "grisignao": "grisignano di zocco",
        "grisignano": "grisignano di zocco",
        "grisignano di zocco": "grisignano di zocco",
        "tezze": "tezze sul brenta",
        "romano": "romano d'ezzelino",
        "bolzano": "bolzano vicentino",
        "rossano": "rossano veneto",
        "torre": "torrebelvicino",
    }

    for year in years:
        file_name = f'Pipistrelli {year}.xlsx'
        if not os.path.exists(file_name): continue

        print(f"\n--- Correzione Geografica v4 {file_name} ---")
        wb = load_workbook(file_name)
        ws = wb.active

        header_row = 1
        col_map = {}
        for r in range(1, 10):
            row_vals = [str(ws.cell(row=r, column=c).value).lower() for c in range(1, ws.max_column + 1)]
            if any('specie' in x or 'comune' in x for x in row_vals):
                header_row = r
                for c in range(1, ws.max_column + 1):
                    val = str(ws.cell(row=r, column=c).value).lower()
                    if 'specie' in val: col_map['specie'] = c
                    if 'comune' in val or 'citt' in val: col_map['comune'] = c
                    if 'prov' in val or val == 'pr': col_map['prov'] = c
                    if 'localit' in val or 'indirizzo' in val: col_map['loc'] = c
                    if 'lat' in val: col_map['lat'] = c
                    if 'lon' in val or 'lng' in val: col_map['lon'] = c
                break

        if 'comune' not in col_map: continue

        count_name = 0
        count_geo = 0
        for r in range(header_row + 1, ws.max_row + 1):
            comune_raw = str(ws.cell(row=r, column=col_map['comune']).value or "").strip().lower()
            prov_raw = str(ws.cell(row=r, column=col_map.get('prov', 1)).value or "").strip().upper()

            # 1. Normalizzazione Nome Comune
            comune_target = comune_raw
            if comune_raw in CORREZIONI_NOMI:
                comune_target = CORREZIONI_NOMI[comune_raw]
            elif "vicenza" in comune_raw or prov_raw == "VI":
                for k, v in CORREZIONI_NOMI.items():
                    if k in comune_raw and len(k) > 3:
                        comune_target = v
                        break

            # 2. Match con Database
            best_info = None
            if comune_target in db_comuni:
                best_info = db_comuni[comune_target]

            if best_info:
                # Applica cambio nome se diverso
                if ws.cell(row=r, column=col_map['comune']).value != comune_target.title():
                    ws.cell(row=r, column=col_map['comune']).value = comune_target.title()
                    count_name += 1

                # Forza provincia
                if 'prov' in col_map and ws.cell(row=r, column=col_map['prov']).value != best_info['prov'].upper():
                    ws.cell(row=r, column=col_map['prov']).value = best_info['prov'].upper()

                # 3. Verifica Geografica
                lat_val = ws.cell(row=r, column=col_map['lat']).value
                lon_val = ws.cell(row=r, column=col_map['lon']).value

                needs_geo_fix = False
                try:
                    curr_lat = float(str(lat_val).replace(',', '.'))
                    curr_lon = float(str(lon_val).replace(',', '.'))
                    dist = get_distance(curr_lat, curr_lon, best_info['lat'], best_info['lon'])
                    if dist > 0.4 or curr_lat < 35 or curr_lat > 48:
                        needs_geo_fix = True
                except:
                    needs_geo_fix = True

                if needs_geo_fix:
                    ws.cell(row=r, column=col_map['lat']).value = best_info['lat']
                    ws.cell(row=r, column=col_map['lon']).value = best_info['lon']
                    count_geo += 1

        wb.save(file_name)
        shutil.copy(file_name, f'app/src/main/assets/{file_name}')
        if os.path.exists('web'): shutil.copy(file_name, f'web/{file_name}')
        print(f"Completato {file_name}: {count_name} nomi corretti, {count_geo} geo-fix.")

if __name__ == "__main__":
    fix_excel_files()
